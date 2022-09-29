import openpyxl as opx

class ExTable():
    arquivo = opx.Workbook()

    def criarTabela(self, nome, conteudo):
        tabela = self.arquivo.create_sheet(nome)

        for y in range(len(conteudo)):

            for x in range(len(conteudo[0])):

                cel = tabela.cell(row=(y+1), column=(x+1))
                # x is a horizontal value
                cel.value = conteudo[y][x]
            
    def saveArchive(self, nome):
        self.arquivo.save(filename=nome)
    